import openpyxl

wb= openpyxl.load_workbook('C:/Users/aneesh goel/Desktop/PythonProjects/Size_FR/abc.xlsx')
sheet=wb.active
Value=[]
for i in range (1,sheet.max_row+1):
     Value.append(sheet.cell(row=i,column=1).value)
print Value
for i in range(0,len(Value)):
    Value[i]=Value[i][Value[i].index('>')+1:]
print Value
