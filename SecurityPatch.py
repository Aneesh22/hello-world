import csv
import openpyxl
#from CreateFile import month1
from openpyxl.styles import Font,Style,PatternFill,Fill,Color,Alignment,Border,Side

#--------------------------------Fetching Data from CSV File------------------------------------#
with open('C:/Users/aneesh goel/Desktop/PythonProjects/Security Patches/MSUDetail.csv','r') as f:
    MSUDetails = csv.reader(f)
    MSUDetailsrows = list(MSUDetails)
    KBNos=[]
    Description=[]
    PackageName=[]
    PackageType=[]
    for row in MSUDetailsrows[1:]:
        for data in row[4:5]:
            OS=data
        for TYPE in row[3:4]:
            PackageType.append(TYPE)
        for KB in row[0:1]:
            KBNos.append('KB'+KB)
            PackageName.append('KB'+KB+'_'+OS+'.exe')
        for Desc in row[5:6]:
            Description.append(Desc)

print OS
print KBNos
print PackageName
print PackageType
ArrayLen=len(PackageName)
for i in range(0,ArrayLen):
    if PackageType[i]!='':
        PackageName[i]=PackageType[i]+'_'+PackageName[i]
print PackageName

#--------------------------------Creating Sheet---------------------------------------------------#
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

Filename=open('C:/Users/aneesh goel/Desktop/PythonProjects/Security Patches/sheetname.txt','r+')
Name=Filename.readlines()
print Name[0]
wb=openpyxl.load_workbook(Name[0])
sheets=wb.get_sheet_names()
print sheets
if OS=='WES7SP1':
    OS='WES7'

#-----------------------------Adding Data to existing sheet-------------------------------#
print 'Verify is the sheet already exists...'
if sheets.__contains__(OS):
    print 'Sheet already Exists'
    sheet=wb.get_sheet_by_name(OS)
    TotalKB=len(KBNos)
    row_count=sheet.max_row
    print 'Current Sheet:',sheet,'\nTotalKB:',TotalKB

    for j in range(row_count+1,row_count+TotalKB+1):
        sheet.cell(row=j,column=1).value=KBNos[(j-1)-row_count]
        sheet.cell(row=j,column=2).value=Description[(j-1)-row_count]
        sheet.cell(row=j,column=3).value=PackageName[(j-1)-row_count]
    for i in range(row_count+1,row_count+TotalKB+1):
        for j in range(1,6):
            sheet.cell(row=i, column=j).border = thin_border
    wb.save(Name[0])
else:
    print 'New Sheet Created '+OS
    SheetIndex=len(sheets)
    wb.create_sheet(index=SheetIndex-1, title=OS)
    sheet=wb.get_sheet_by_name(OS)
    ListFormat=['KB No.','Description','WDM Package Name','Affected software','Dependency']
    sheet['A1']='Security Patches JAN 2016-'+OS+''

#------------------------Fonts and Styling------------------------------------------#
    font1 = Font(bold=True)
    BackColor=PatternFill(patternType='solid', fgColor=Color('FFFF00'))
    BackColor2=PatternFill(patternType='solid', fgColor=Color('528ED5'))
    align=Alignment(horizontal='center')
    sheet.merge_cells('A1:E1')
    for i in range(1,6):
        sheet.cell(row=1,column=i).style=Style(alignment=align,fill=BackColor2,font=font1)
        sheet.cell(row=2,column=i).style=Style(fill=BackColor,font=font1)
        sheet.cell(row=2,column=i).value=ListFormat[i-1]

#-----------------------------Adding Data to new sheet-------------------------------#
    TotalKB=len(KBNos)
    for j in range(3,TotalKB+3):
        sheet.cell(row=j,column=1).value=KBNos[j-3]
        sheet.cell(row=j,column=2).value=Description[j-3]
        sheet.cell(row=j,column=3).value=PackageName[j-3]
    for i in range(1,TotalKB+3):
        for j in range(1,6):
            sheet.cell(row=i, column=j).border = thin_border
    wb.save(Name[0])
