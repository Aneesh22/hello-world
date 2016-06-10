import openpyxl
from openpyxl.styles import *
#-------------------------------Reading Text file and Extracting Data------------------#
fil=open('C:/Users/aneesh goel/Desktop/PythonProjects/Size_FR/wes9out_FR1.txt','r')
fil1=open('C:/Users/aneesh goel/Desktop/PythonProjects/Size_FR/wes9out_FR2.txt','r')
fil2=open('C:/Users/aneesh goel/Desktop/PythonProjects/Size_FR/wes9out_FR3.txt','r')

#-------------------------------Reading WFR1 Text file and Extracting Data------------------#
lines=fil.readlines()
SplitArray,PackageSize,WFR1Status=[[]]*3
KbNumber=[]
TotalSize=[]
for line in lines:
    if '-after' in line:
        SplitArray=line.split(',')
        TotalSize.append(int(SplitArray[5][2:len(SplitArray[5])-2]))
        KbNumber.append(SplitArray[1])
        if SplitArray[3].__contains__('fail'):
            WFR1Status.append('NA')
        else:
            WFR1Status.append('PASS')
print KbNumber
print WFR1Status
print TotalSize
fil.close()

#-------------------------------Reading WFR2 Text file and Extracting Data------------------#
lines=fil1.readlines()
SplitArray1=WFR2Status=[]
TotalSize1=[]
for line in lines:
    if '-after' in line:
        SplitArray1=line.split(',')
        TotalSize1.append(int(SplitArray1[5][2:len(SplitArray1[5])-2]))
        if SplitArray1[3].__contains__('fail'):
            WFR2Status.append('NA')
        else:
            WFR2Status.append('PASS')
print WFR2Status
print TotalSize1
fil1.close()
#-------------------------------Reading WFR3 Text file and Extracting Data------------------#
lines=fil2.readlines()
SplitArray2=WFR3Status=[]
TotalSize2=[]
for line in lines:
    if '-after,' in line:
        SplitArray2=line.split(',')
        TotalSize2.append(int(SplitArray2[5][2:len(SplitArray2[5])-2]))
        if SplitArray2[3].__contains__('fail'):
            WFR3Status.append('NA')
        else:
            WFR3Status.append('PASS')
print WFR3Status
print TotalSize2
fil1.close()

#------------------------------Writing in the Excel file--------------------------------#
FinalSize=[]
for x in range(0,len(TotalSize)):
    FinalSize.append(max(int(TotalSize[x]),int(TotalSize1[x]),int(TotalSize2[x])))
    TotalSize[x]=int((FinalSize[x]))
    TotalSize[x]=TotalSize[x]/1024
    if TotalSize[x] in range(0,5):
        TotalSize[x]=10
    elif TotalSize[x] in range(5,10):
        TotalSize[x]=15
    elif TotalSize[x] in range(10,15):
        TotalSize[x]=20
    else:
        TotalSize[x]=TotalSize[x]+10
print TotalSize

font1 = Font(bold=True)
BackColor=PatternFill(patternType='solid', fgColor=Color('FFFF00'))

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

FileName='C:/Users/aneesh goel/Desktop/PythonProjects/Size_FR/SizeSheet.xlsx'
Format=['KB NUMBER','WFR1','WFR2','WFR3','SIZE']
wb= openpyxl.Workbook()
sheet=wb.get_active_sheet()
for i in range (1,6):
     sheet.cell(row=1,column=i).value=Format[i-1]
     sheet.cell(row=1,column=i).style=Style(fill=BackColor,font=font1)
TotalKb=len(KbNumber)
array=[KbNumber,WFR1Status,WFR2Status,WFR3Status,TotalSize]
for i in range(0,len(array)):
    for j in range (0,TotalKb):
        sheet.cell(row=j+2,column=i+1).value=array[i][j]

for i in range(0,len(array)):
    for j in range (0,TotalKb+1):
        sheet.cell(row=j+1, column=i+1).border = thin_border

wb.save(FileName)